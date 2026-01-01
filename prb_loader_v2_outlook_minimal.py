import re
import win32com.client
from openpyxl import load_workbook

# -------- CONFIG --------
BASE_DIR = r"E:\Jazz\Automation\MW Tracker Email Automation"

EXCEL_FILE = rf"{BASE_DIR}\MW_Tickets_Raw.xlsx"
SHEET_NAME = "Sheet1"
PRB_COLUMN = 1  # Column A

MAILBOX_NAME = "ahmed.mujtaba@jazz.com.pk"
TARGET_FOLDER = "PRB TT"

# -------- REGEX --------
PRB_REGEX = re.compile(r"(PRB\d{7})", re.IGNORECASE)

# -------- LOAD EXCEL --------
wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

existing_prbs = {
    ws.cell(row=r, column=PRB_COLUMN).value
    for r in range(2, ws.max_row + 1)
    if ws.cell(row=r, column=PRB_COLUMN).value
}

# -------- CONNECT TO OUTLOOK --------
outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

root = namespace.Folders[MAILBOX_NAME]
prb_folder = root.Folders[TARGET_FOLDER]
messages = prb_folder.Items

# Optional: newest emails first
messages.Sort("[ReceivedTime]", True)

# -------- PROCESS EMAILS --------
for msg in messages:
    try:
        subject = msg.Subject or ""
    except Exception:
        continue

    subject_lower = subject.lower()

    # شرط: must contain BOTH words
    if "high" not in subject_lower or "utilization" not in subject_lower:
        continue

    match = PRB_REGEX.search(subject)
    if not match:
        continue

    prb = match.group(1).upper()

    if prb in existing_prbs:
        continue

    ws.cell(row=ws.max_row + 1, column=PRB_COLUMN).value = prb
    existing_prbs.add(prb)

    print(f"✅ Added {prb}")

# -------- SAVE --------
wb.save(EXCEL_FILE)
print("✔ Done. Outlook PRBs synced safely.")
