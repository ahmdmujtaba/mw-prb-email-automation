import os
import re
import extract_msg
from openpyxl import load_workbook

# -------- CONFIG --------
BASE_DIR = r"E:\Jazz\Automation\MW Tracker Email Automation"

EXCEL_FILE = rf"{BASE_DIR}\MW_Tickets_Raw.xlsx"
EMAIL_DIR = rf"{BASE_DIR}\emails"

SHEET_NAME = "Sheet1"
PRB_COLUMN = 1  # Column A

# -------- REGEX --------
PRB_REGEX = re.compile(r"(PRB\d{7})", re.IGNORECASE)

# -------- LOAD EXCEL --------
wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

# Read existing PRBs from column A
existing_prbs = {
    ws.cell(row=r, column=PRB_COLUMN).value
    for r in range(2, ws.max_row + 1)
    if ws.cell(row=r, column=PRB_COLUMN).value
}

# -------- PROCESS EMAILS --------
for file in os.listdir(EMAIL_DIR):
    if not file.lower().endswith(".msg"):
        continue

    msg = extract_msg.Message(os.path.join(EMAIL_DIR, file))
    subject = msg.subject or ""

    match = PRB_REGEX.search(subject)
    if not match:
        print(f"❌ No PRB found in subject: {file}")
        continue

    prb = match.group(1).upper()

    if prb in existing_prbs:
        print(f"⏭ PRB already exists: {prb}")
        continue

    # Append PRB to last unused row
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=PRB_COLUMN).value = prb
    existing_prbs.add(prb)

    print(f"✅ Added {prb}")

# -------- SAVE --------
wb.save(EXCEL_FILE)
print("✔ Done. PRBs synced safely.")
